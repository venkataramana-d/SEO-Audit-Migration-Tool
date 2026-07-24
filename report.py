"""
Report — builds the downloadable Excel workbook from audit results.

Sheets:
  Summary        — headline metrics + issue breakdown
  All Pages      — one row per URL with SEO fields + verdict (color-coded)
  Issues         — one row per issue: severity, location, how to fix, why
  Redirect Map   — old URL -> [new URL] template for the migration
  Priority Pages — GSC top performers to preserve (if GSC data supplied)
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD_FILL = PatternFill("solid", fgColor="1A237E")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
CLEAR_FILL = PatternFill("solid", fgColor="C8E6C9")
ISSUE_FILL = PatternFill("solid", fgColor="FFCDD2")
SEV_FILL = {
    "Critical": PatternFill("solid", fgColor="C62828"),
    "High": PatternFill("solid", fgColor="EF6C00"),
    "Medium": PatternFill("solid", fgColor="F9A825"),
    "Low": PatternFill("solid", fgColor="90A4AE"),
}
SEV_FONT = {
    "Critical": Font(bold=True, color="FFFFFF"),
    "High": Font(bold=True, color="FFFFFF"),
    "Medium": Font(bold=True, color="000000"),
    "Low": Font(bold=True, color="FFFFFF"),
}
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, wrap_cols=(), max_w=70, sample=300):
    # size columns from the header + a sample of rows (scanning every cell on a
    # 60k-row sheet is what made large reports hang)
    for col in ws.iter_cols(max_row=min(ws.max_row, sample)):
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(width + 3, max_w)
    for wc in wrap_cols:
        ws.column_dimensions[wc].width = max_w
        for cell in ws[wc][:sample]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_workbook(summary, page_index, all_issues, meta, gsc_rows=None):
    wb = Workbook()

    # ---------- Summary ----------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Pre-Migration SEO Audit"
    ws["A1"].font = Font(bold=True, size=16, color="1A237E")
    ws["A3"] = "Property"
    ws["B3"] = meta.get("start_url", "")
    ws["A4"] = "Date"
    ws["B4"] = meta.get("date", "")
    ws["A5"] = "Pages crawled"
    ws["B5"] = summary["pages_crawled"]
    ws["A6"] = "Pages fully clear"
    ws["B6"] = summary["pages_clear"]
    ws["A7"] = "Pages with issues"
    ws["B7"] = summary["pages_with_issues"]
    ws["A8"] = "Total issues found"
    ws["B8"] = summary["total_issues"]
    for r in range(3, 9):
        ws[f"A{r}"].font = Font(bold=True)

    ws["A10"] = "Issues by severity"
    ws["A10"].font = Font(bold=True, size=12)
    ws.append([])
    row = 11
    ws[f"A{row}"] = "Severity"; ws[f"B{row}"] = "Count"
    ws[f"A{row}"].fill = HEAD_FILL; ws[f"A{row}"].font = HEAD_FONT
    ws[f"B{row}"].fill = HEAD_FILL; ws[f"B{row}"].font = HEAD_FONT
    for sev in ["Critical", "High", "Medium", "Low"]:
        row += 1
        ws[f"A{row}"] = sev
        ws[f"B{row}"] = summary["by_severity"].get(sev, 0)
        ws[f"A{row}"].fill = SEV_FILL[sev]
        ws[f"A{row}"].font = SEV_FONT[sev]

    row += 2
    ws[f"A{row}"] = "Issues by type"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = "Check"; ws[f"B{row}"] = "Count"
    ws[f"A{row}"].fill = HEAD_FILL; ws[f"A{row}"].font = HEAD_FONT
    ws[f"B{row}"].fill = HEAD_FILL; ws[f"B{row}"].font = HEAD_FONT
    for check, cnt in summary["by_check"].items():
        row += 1
        ws[f"A{row}"] = check
        ws[f"B{row}"] = cnt
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60

    # ---------- All Pages ----------
    ws = wb.create_sheet("All Pages")
    headers = ["URL", "Status", "Verdict", "Issues", "Title", "Title Len",
               "Desc Len", "H1s", "Words", "Imgs no ALT", "Int Links",
               "Ext Links", "Resp ms", "Canonical", "Robots"]
    _header(ws, headers)
    for p in page_index:
        ws.append([
            p["url"], p["status"], p["verdict"], p["issue_count"], p["title"],
            p["title_len"], p["meta_desc_len"], p["h1_count"], p["word_count"],
            p["img_missing_alt"], p["internal_links"], p["external_links"],
            p["response_ms"], p["canonical"], p["meta_robots"],
        ])
        vcell = ws.cell(row=ws.max_row, column=3)
        vcell.fill = CLEAR_FILL if p["verdict"] == "CLEAR" else ISSUE_FILL
        vcell.font = Font(bold=True)
    _autosize(ws, wrap_cols=())
    ws.auto_filter.ref = ws.dimensions

    # ---------- Issues ----------
    ws = wb.create_sheet("Issues")
    headers = ["Severity", "Check", "URL (location)", "Element", "Detail",
               "How to fix", "Why it matters"]
    _header(ws, headers)
    sev_rank = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    for i in sorted(all_issues, key=lambda x: sev_rank.get(x["severity"], 9)):
        ws.append([
            i["severity"], i["check"], i["url"], i["location"],
            i["detail"], i["how_to_fix"], i["why"],
        ])
        scell = ws.cell(row=ws.max_row, column=1)
        scell.fill = SEV_FILL.get(i["severity"], SEV_FILL["Low"])
        scell.font = SEV_FONT.get(i["severity"], SEV_FONT["Low"])
    _autosize(ws, wrap_cols=("E", "F", "G"))
    ws.column_dimensions["C"].width = 55
    ws.auto_filter.ref = ws.dimensions

    # ---------- Redirect Map ----------
    ws = wb.create_sheet("Redirect Map")
    _header(ws, ["Old URL (current site)", "New URL (fill in)", "Status", "Notes"])
    for p in page_index:
        ws.append([p["url"], "", p["status"], ""])
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["D"].width = 40
    ws.auto_filter.ref = ws.dimensions

    # ---------- Priority Pages (GSC) ----------
    if gsc_rows:
        ws = wb.create_sheet("Priority Pages (GSC)")
        _header(ws, ["Page", "Clicks", "Impressions", "CTR %", "Avg Position"])
        for r in gsc_rows:
            ws.append(r)
        ws.column_dimensions["A"].width = 60
        ws.auto_filter.ref = ws.dimensions

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
