"""
SEO Pre-Migration Audit Tool — local web app.

Run:  python app.py
Open: http://127.0.0.1:5000

Configure the target path, page limit, and options in the UI, run the audit,
watch live progress, then download the color-coded Excel report.
"""

import threading
import uuid
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file, abort

import json
import os

import crawler
import auditor
import report
import gsc
import perf
import redirects

# absolute paths so templates resolve regardless of the working directory
# (needed when deployed under a WSGI host such as Vercel/gunicorn)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"))
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True

# in-memory job store: job_id -> state dict
JOBS = {}
CLIENT_FILE = os.path.join(BASE_DIR, "..", "oauth_client.json")
TOKEN_FILE = os.path.join(BASE_DIR, "..", "token.json")
# drop a Google service-account key here to connect GSC permanently (no browser,
# no expiry). Takes priority over OAuth when present.
SERVICE_ACCOUNT_FILE = os.path.join(BASE_DIR, "..", "service_account.json")
SETTINGS_FILE = os.path.join(BASE_DIR, "settings.json")


def load_settings():
    try:
        with open(SETTINGS_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, ValueError, OSError):
        return {}


def save_settings(data):
    try:
        with open(SETTINGS_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except OSError:
        pass  # read-only filesystem (e.g. serverless host) — ignore


def run_job(job_id, params):
    job = JOBS[job_id]
    try:
        stop = job["stop_flag"]

        def progress(**kw):
            job.update(kw)

        # resolve the URL list depending on input mode
        url_list = params.get("url_list")
        if params.get("mode") == "sitemap" and params.get("sitemap_url"):
            job["phase"] = "sitemap"
            url_list = crawler.collect_sitemap_urls(params["sitemap_url"],
                                                    max_urls=params["max_pages"])
            if not url_list:
                job["error"] = "No URLs found in that sitemap."
                job["phase"] = "error"
                job["done"] = True
                return

        job["phase"] = "crawling"
        c = crawler.Crawler(
            start_url=params["start_url"],
            max_pages=params["max_pages"],
            workers=params["workers"],
            progress_cb=progress,
            stop_flag=stop,
            url_list=url_list,
        )
        pages, status_map = c.run()
        if stop.is_set():
            job["phase"] = "stopped"
            job["done"] = True
            return

        job["phase"] = "auditing"
        all_issues, page_index = auditor.audit_all(pages)
        summary = auditor.summarize(page_index, all_issues)
        summary["total_found"] = len(c.seen)          # URLs discovered / provided
        summary["sitemap_total"] = c.sitemap_total
        # in sitemap mode, report how many URLs the sitemap declared
        summary["sitemap_scope"] = len(url_list) if (params.get("mode") == "sitemap"
                                                     and url_list) else 0
        summary["mode"] = params.get("mode", "url")

        # group issues by URL for the interactive explorer
        from collections import defaultdict
        issues_by_url = defaultdict(list)
        for i in all_issues:
            issues_by_url[i["url"]].append(i)
        job["page_index"] = page_index
        job["issues_by_url"] = issues_by_url
        job["pages_by_url"] = {p.get("url"): p for p in pages}

        def _num(x):
            try:
                return float(str(x).replace(",", "").replace("%", "").strip())
            except (ValueError, TypeError):
                return 0

        def _to_map(rows):
            m = {}
            for rank, r in enumerate(sorted(rows, key=lambda x: _num(x[1]), reverse=True), 1):
                m[str(r[0]).rstrip("/")] = {
                    "clicks": int(_num(r[1])), "impressions": int(_num(r[2])),
                    "ctr": round(_num(r[3]), 2), "position": round(_num(r[4]), 1),
                    "rank": rank,
                }
            return m

        gsc_map_90, gsc_map_365 = {}, {}
        upload = params.get("gsc_upload")      # rows uploaded from a GSC CSV export
        upload_12 = params.get("gsc_upload_12")
        if upload:
            # GSC data supplied by CSV upload — no API/login needed (works anywhere)
            job["phase"] = "gsc"
            try:
                gsc_map_90 = _to_map(upload)
                gsc_map_365 = _to_map(upload_12) if upload_12 else gsc_map_90
            except Exception as e:
                job["gsc_error"] = "uploaded GSC data: " + str(e)[:200]
        elif params.get("include_gsc"):
            job["phase"] = "gsc"
            try:
                from concurrent.futures import ThreadPoolExecutor

                def _pull(days):
                    return gsc.top_pages(
                        site_url=params["gsc_property"], client_file=CLIENT_FILE,
                        token_file=TOKEN_FILE, days=days,
                        path_filter=params.get("path_filter"),
                        sa_file=SERVICE_ACCOUNT_FILE)

                # fetch the 3-month and 12-month windows concurrently (was ~8s serial)
                with ThreadPoolExecutor(max_workers=2) as gex:
                    f90 = gex.submit(_pull, 90)
                    f365 = gex.submit(_pull, 365)
                    gsc_map_90 = _to_map(f90.result())
                    gsc_map_365 = _to_map(f365.result())
            except Exception as e:
                job["gsc_error"] = str(e)[:300]
        job["gsc_map"] = gsc_map_90          # 3-month is the default (used by Excel too)
        job["gsc_map_90"] = gsc_map_90
        job["gsc_map_365"] = gsc_map_365

        # The Excel report is built ON DEMAND (see /api/download), not here — for
        # large crawls (tens of thousands of issue rows) it can take minutes and
        # was blocking the whole job from completing. Store what it needs instead.
        job["report_meta"] = {"start_url": params["start_url"],
                              "date": datetime.now().strftime("%Y-%m-%d %H:%M")}
        job["summary"] = summary
        job["phase"] = "complete"
        job["done"] = True
    except Exception as e:
        import traceback
        job["error"] = str(e)
        job["traceback"] = traceback.format_exc()
        job["phase"] = "error"
        job["done"] = True


@app.route("/")
def index():
    return render_template("index.html", gsc_available=gsc.available())


@app.route("/checklist")
def checklist():
    return render_template("checklist.html")


@app.route("/meta")
def meta_page():
    return render_template("meta.html")


@app.route("/api/default_urls")
def default_urls():
    """Pre-built list of the site's live blogs & courses (from the sitemap)."""
    try:
        with open(os.path.join(BASE_DIR, "default_urls.json")) as f:
            return jsonify(json.load(f))
    except (FileNotFoundError, ValueError):
        return jsonify({"all": [], "courses": [], "blogs": []})


@app.route("/settings", methods=["GET", "POST"])
def settings():
    s = load_settings()
    saved = False
    if request.method == "POST":
        s["pagespeed_api_key"] = (request.form.get("pagespeed_api_key") or "").strip()
        save_settings(s)
        saved = True
    kind = gsc.connection_kind(CLIENT_FILE, TOKEN_FILE, SERVICE_ACCOUNT_FILE)
    return render_template("settings.html", settings=s, saved=saved,
                           gsc_connected=bool(kind), gsc_kind=kind,
                           sa_path=SERVICE_ACCOUNT_FILE,
                           client_exists=os.path.exists(CLIENT_FILE))


@app.route("/page")
def page_detail():
    job_id = request.args.get("job", "")
    url = request.args.get("url", "")
    job = JOBS.get(job_id)
    if not job or not job.get("pages_by_url"):
        return render_template("detail.html", not_found=True, url=url), 404
    page = job["pages_by_url"].get(url)
    if not page:
        return render_template("detail.html", not_found=True, url=url), 404
    issues = job.get("issues_by_url", {}).get(url, [])
    sev_rank = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    issues = sorted(issues, key=lambda i: sev_rank.get(i["severity"], 9))
    gsc_stats = job.get("gsc_map", {}).get(url.rstrip("/"))
    return render_template("detail.html", not_found=False, job_id=job_id,
                           page=page, issues=issues, gsc=gsc_stats,
                           canon_status=auditor.canonical_status(page),
                           has_perf_key=bool(load_settings().get("pagespeed_api_key")))


@app.route("/page/links")
def page_links():
    job_id = request.args.get("job", "")
    url = request.args.get("url", "")
    job = JOBS.get(job_id)
    if not job or not job.get("pages_by_url") or url not in job["pages_by_url"]:
        return render_template("link_analysis.html", not_found=True, url=url), 404
    page = job["pages_by_url"][url]
    links = []
    for l in page.get("links", []):
        st = l.get("status")
        issue = None
        if st is not None and (st == 0 or st >= 400):
            issue = "broken"
        elif l["internal"] and l["nofollow"]:
            issue = "internal-nofollow"
        elif (not l["internal"]) and (not l["nofollow"]) \
                and not l.get("sponsored") and not l.get("ugc"):
            issue = "external-dofollow"
        elif l["target_blank"] and not l["noopener"]:
            issue = "unsafe-blank"
        links.append({**l, "issue": issue})
    return render_template("link_analysis.html", not_found=False, job_id=job_id,
                           page=page, links=links, url=url)


@app.route("/api/perf")
def api_perf():
    url = request.args.get("url", "")
    strategy = request.args.get("strategy", "mobile")
    if not url.startswith("http"):
        return jsonify({"error": "invalid url"}), 400
    key = load_settings().get("pagespeed_api_key", "")
    return jsonify(perf.pagespeed(url, api_key=key, strategy=strategy))


@app.route("/api/start", methods=["POST"])
def start():
    from urllib.parse import urlparse
    data = request.get_json(force=True)
    mode = (data.get("input_mode") or "url").lower()

    # collect the list of URLs to audit depending on the input mode
    url_list = None
    sitemap_url = None
    if mode in ("csv", "paste"):
        url_list = [u.strip() for u in (data.get("url_list") or []) if u.strip().startswith("http")]
        if not url_list:
            return jsonify({"error": "No valid http(s) URLs found in the uploaded/pasted list."}), 400
        base = url_list[0]
    elif mode == "sitemap":
        sitemap_url = (data.get("sitemap_url") or "").strip()
        if not sitemap_url.startswith("http"):
            return jsonify({"error": "Enter a valid sitemap URL (https://…/sitemap.xml)."}), 400
        base = sitemap_url
    else:  # url crawl
        base = (data.get("start_url") or "").strip()
        if not base.startswith("http"):
            return jsonify({"error": "Enter a valid URL starting with http/https."}), 400

    try:
        max_pages = max(1, min(int(data.get("max_pages", 500)), 15000))
    except (TypeError, ValueError):
        max_pages = 500

    pu = urlparse(base)
    origin = f"{pu.scheme}://{pu.netloc}/"
    # in crawl mode we scope by the URL's path; list/sitemap modes audit all given URLs
    scope_path = pu.path if (mode == "url" and pu.path and pu.path != "/") else None

    job_id = uuid.uuid4().hex[:12]
    JOBS[job_id] = {
        "phase": "starting", "crawled": 0, "found": 0, "current": "",
        "done": False, "error": None, "stop_flag": threading.Event(),
    }
    params = {
        "mode": mode,
        "start_url": base if mode == "url" else origin,
        "url_list": url_list,
        "sitemap_url": sitemap_url,
        "max_pages": max_pages,
        "workers": int(data.get("workers", 16)),
        "include_gsc": bool(data.get("include_gsc")),
        "gsc_upload": data.get("gsc_upload"),       # rows from an uploaded GSC CSV
        "gsc_upload_12": data.get("gsc_upload_12"),
        "gsc_property": origin,
        "path_filter": scope_path,
    }
    t = threading.Thread(target=run_job, args=(job_id, params), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/progress/<job_id>")
def progress(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "unknown job"}), 404
    return jsonify({
        "phase": job.get("phase"),
        "crawled": job.get("crawled", 0),
        "found": job.get("found", 0),
        "current": job.get("current", ""),
        "done": job.get("done", False),
        "error": job.get("error"),
        "gsc_error": job.get("gsc_error"),
        "summary": job.get("summary"),
        "has_report": bool(job.get("report_bytes")),
    })


@app.route("/api/results/<job_id>")
def results(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "unknown job"}), 404
    return jsonify({
        "summary": job.get("summary"),
        "pages": job.get("page_index", []),
        "issues_by_url": job.get("issues_by_url", {}),
        "gsc_map": job.get("gsc_map", {}),
        "gsc_map_90": job.get("gsc_map_90", {}),
        "gsc_map_365": job.get("gsc_map_365", {}),
    })


@app.route("/api/stop/<job_id>", methods=["POST"])
def stop(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "unknown job"}), 404
    job["stop_flag"].set()
    return jsonify({"ok": True})


@app.route("/api/download/<job_id>")
def download(job_id):
    job = JOBS.get(job_id)
    if not job or not job.get("page_index"):
        abort(404)
    # build the workbook on demand (and cache it on the job)
    if not job.get("report_bytes"):
        page_index = job["page_index"]
        all_issues = [i for lst in job.get("issues_by_url", {}).values() for i in lst]
        gsc_map = job.get("gsc_map_90", {})
        gsc_rows = [[u, v["clicks"], v["impressions"], v["ctr"], v["position"]]
                    for u, v in sorted(gsc_map.items(),
                                       key=lambda x: -x[1]["clicks"])] or None
        bio = report.build_workbook(job["summary"], page_index, all_issues,
                                    job.get("report_meta", {}), gsc_rows)
        job["report_bytes"] = bio.getvalue()
    from io import BytesIO
    bio = BytesIO(job["report_bytes"])
    bio.seek(0)
    fname = f"SEO-Audit-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return send_file(
        bio, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ------------------------- Redirect Validation -------------------------

def _validate_pairs(job, pairs):
    from concurrent.futures import ThreadPoolExecutor, as_completed
    stop = job["stop_flag"]
    job.update(phase="validating", crawled=0, found=len(pairs))
    session = redirects.make_session()
    collected = []                      # streamed live to the UI as each completes
    job["redirect_results"] = collected
    job["redirect_summary"] = redirects.summarize(collected)
    done = 0
    # deliberately gentle concurrency — redirect checks do multiple GETs each and
    # bot-protected sites (Cloudflare/WAF) throttle bursts, returning connection
    # errors or slow challenge pages that look like false failures
    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {ex.submit(redirects.validate, p["source"], p.get("expected", ""), session): i
                for i, p in enumerate(pairs)}
        for f in as_completed(futs):
            if stop.is_set():
                break
            i = futs[f]
            try:
                r = f.result()
            except Exception as e:
                r = {"source": pairs[i]["source"], "expected": pairs[i].get("expected", ""),
                     "final_url": pairs[i]["source"], "final_status": 0,
                     "redirect_type": None, "hops": 0, "chain": [], "is_chain": False,
                     "is_loop": False, "expected_match": None, "canonical": "",
                     "canonical_ok": True, "indexable": True, "index_reason": "",
                     "is_https": False, "www": {}, "trailing_slash": {}, "lost_params": [],
                     "lost_utm": [], "speed_ms": 0, "slow": False,
                     "issues": [{"severity": "Critical", "msg": "Validation error: " + str(e)[:120]}],
                     "result": "FAIL"}
            collected.append(r)
            done += 1
            if done % 8 == 0 or done == len(pairs):
                job["redirect_summary"] = redirects.summarize(collected)
            job.update(phase="validating", crawled=done, found=len(pairs),
                       current=pairs[i]["source"])
    job["redirect_results"] = collected
    job["redirect_summary"] = redirects.summarize(collected)
    job["phase"] = "complete"
    job["done"] = True


def run_redirects(job_id, pairs):
    job = JOBS[job_id]
    try:
        _validate_pairs(job, pairs)
    except Exception as e:
        import traceback
        job["error"] = str(e)
        job["traceback"] = traceback.format_exc()
        job["phase"] = "error"
        job["done"] = True


def run_redirects_crawl(job_id, start_url, max_pages):
    """Crawl a site to discover URLs, then validate redirects on each."""
    job = JOBS[job_id]
    try:
        job["phase"] = "crawling"
        c = crawler.Crawler(start_url=start_url, max_pages=max_pages, workers=16,
                            progress_cb=lambda **k: job.update(k), stop_flag=job["stop_flag"])
        pages, _ = c.run()
        if job["stop_flag"].is_set():
            job["phase"] = "stopped"
            job["done"] = True
            return
        # de-duplicate the discovered URLs
        seen, urls = set(), []
        for p in pages:
            u = p.get("url")
            if u and u not in seen:
                seen.add(u)
                urls.append(u)
        _validate_pairs(job, [{"source": u, "expected": ""} for u in urls])
    except Exception as e:
        import traceback
        job["error"] = str(e)
        job["traceback"] = traceback.format_exc()
        job["phase"] = "error"
        job["done"] = True


@app.route("/redirects")
def redirects_page():
    return render_template("redirects.html")


@app.route("/api/redirects/start", methods=["POST"])
def redirects_start():
    data = request.get_json(force=True)
    # crawl-then-validate mode
    crawl = data.get("crawl")
    if crawl:
        start_url = (crawl.get("start_url") or "").strip()
        if not start_url.startswith("http"):
            return jsonify({"error": "Enter a valid website URL to crawl."}), 400
        try:
            max_pages = max(1, min(int(crawl.get("max_pages", 500)), 15000))
        except (TypeError, ValueError):
            max_pages = 500
        job_id = uuid.uuid4().hex[:12]
        JOBS[job_id] = {"phase": "starting", "crawled": 0, "found": 0, "current": "",
                        "done": False, "error": None, "stop_flag": threading.Event()}
        threading.Thread(target=run_redirects_crawl,
                         args=(job_id, start_url, max_pages), daemon=True).start()
        return jsonify({"job_id": job_id})

    raw = data.get("pairs") or []
    pairs = [{"source": (p.get("source") or "").strip(),
              "expected": (p.get("expected") or "").strip()}
             for p in raw if (p.get("source") or "").strip().startswith("http")]
    if not pairs:
        return jsonify({"error": "No valid source URLs (must start with http/https)."}), 400
    pairs = pairs[:15000]
    job_id = uuid.uuid4().hex[:12]
    JOBS[job_id] = {"phase": "starting", "crawled": 0, "found": len(pairs),
                    "current": "", "done": False, "error": None,
                    "stop_flag": threading.Event()}
    threading.Thread(target=run_redirects, args=(job_id, pairs), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/redirects/results/<job_id>")
def redirects_results(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "unknown job"}), 404
    return jsonify({"summary": job.get("redirect_summary"),
                    "results": job.get("redirect_results", [])})


@app.route("/api/redirects/sitemap", methods=["POST"])
def redirects_sitemap():
    data = request.get_json(force=True)
    sm = (data.get("sitemap_url") or "").strip()
    if not sm.startswith("http"):
        return jsonify({"error": "Enter a valid sitemap URL."}), 400
    urls = crawler.collect_sitemap_urls(sm, max_urls=15000)
    return jsonify({"urls": urls})


@app.route("/api/redirects/upload_xlsx", methods=["POST"])
def redirects_upload_xlsx():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded."}), 400
    from openpyxl import load_workbook
    from io import BytesIO
    try:
        wb = load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": "Could not read Excel file: " + str(e)[:120]}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"pairs": []})
    # decide header vs data
    first = rows[0]
    header_is_data = first and first[0] and str(first[0]).strip().lower().startswith("http")
    si, ei = 0, 1
    start = 0
    if not header_is_data:
        hdr = {str(c).lower().strip(): i for i, c in enumerate(first or []) if c}
        def find(*keys):
            for k in keys:
                for h, i in hdr.items():
                    if k in h:
                        return i
            return None
        si = find("source", "url", "from", "old") or 0
        ei = find("expected", "destination", "landing", "target", "new", "to")
        start = 1
    pairs = []
    for r in rows[start:]:
        if not r:
            continue
        src = str(r[si]).strip() if si < len(r) and r[si] else ""
        exp = ""
        if ei is not None and ei < len(r) and r[ei]:
            exp = str(r[ei]).strip()
        if src.startswith("http"):
            pairs.append({"source": src, "expected": exp})
    return jsonify({"pairs": pairs})


@app.route("/api/redirects/export/<job_id>")
def redirects_export(job_id):
    job = JOBS.get(job_id)
    if not job or not job.get("redirect_results"):
        abort(404)
    results = job["redirect_results"]
    fmt = request.args.get("fmt", "csv")
    from io import BytesIO, StringIO
    cols = ["Source URL", "Expected", "Actual Final URL", "Result", "Redirect Type",
            "Hops", "Chain", "Loop", "Final Status", "Canonical OK", "Indexable",
            "HTTPS", "Speed (ms)", "Issues"]

    def row_of(r):
        return [r["source"], r.get("expected", ""), r["final_url"], r["result"],
                r.get("redirect_type"), r["hops"], "yes" if r["is_chain"] else "no",
                "yes" if r["is_loop"] else "no", r["final_status"],
                "yes" if r["canonical_ok"] else "no", "yes" if r["indexable"] else "no",
                "yes" if r["is_https"] else "no", r["speed_ms"],
                " | ".join(i["msg"] for i in r["issues"])]

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Redirects"
        ws.append(cols)
        for r in results:
            ws.append(row_of(r))
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True,
                         download_name=f"Redirect-Validation-{datetime.now():%Y%m%d-%H%M}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    import csv
    sio = StringIO()
    w = csv.writer(sio)
    w.writerow(cols)
    for r in results:
        w.writerow(row_of(r))
    bio = BytesIO(sio.getvalue().encode("utf-8-sig"))
    return send_file(bio, as_attachment=True,
                     download_name=f"Redirect-Validation-{datetime.now():%Y%m%d-%H%M}.csv",
                     mimetype="text/csv")


if __name__ == "__main__":
    print("\n  SEO Pre-Migration Audit Tool")
    print("  Open http://127.0.0.1:5000 in your browser\n")
    app.run(host="127.0.0.1", port=5000, debug=False)
